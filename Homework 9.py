import sys
import json
import os
from PyQt6.QtWidgets import QApplication, QWidget, QLabel, QLineEdit, QPushButton, QVBoxLayout, QTabWidget, QTextBrowser, QMessageBox
from PyQt6.QtCore import Qt
from openpyxl import Workbook
 
class Student:
    def __init__(self, student_id, name):
        self.student_id = student_id
        self.name = name
        self.courses = {}
 
    def add_course(self, course_code, course_name, semester):
        if semester not in self.courses:
            self.courses[semester] = []
            self.courses[semester].append({"code": course_code, "name": course_name})
 
    def query_course_info(self, semester):
        if semester in self.courses:
            courses_in_semester = self.courses[semester]
            return f"Courses taken by {self.name} in {semester}:", courses_in_semester
        else:
            return f"No courses found for {self.name} in semester {semester}.", None
 
    def get_all_courses_info(self):
        all_courses_info = []
        for semester, courses in self.courses.items():
            for course in courses:
                all_courses_info.append({"student_id": self.student_id,
                                         "name": self.name,
                                         "semester": semester,
                                         "course_code": course["code"],
                                         "course_name": course["name"]
                                        })
                return all_courses_info
 
    def save_to_file(self, filename):
        with open(filename, 'w') as file:
            data = {
                    "student_id": self.student_id,
                    "name": self.name,
                    "courses": self.courses
                   }
            json.dump(data, file)
 
    def load_from_file(self, filename):
        if os.path.exists(filename):
            with open(filename, 'r') as file:
                data = json.load(file)
                self.student_id = data["student_id"]
                self.name = data["name"]
                self.courses = data["courses"]
                
    def generate_sample_data():
        sample_students = []
        for i in range(10):
            student_id = f"STUST{i+1:03d}"
            name = f"Student{i+1}"
            student = Student(student_id, name)
            for course_code, course_name in [("Python", "Python Programming"), ("Java", "Java Programming"),
                                             ("C++", "C++ Programming"), ("JavaScript", "JavaScript Programming"),
                                             ("Database", "Database Management"), ("OS", "Operating Systems")]:
                 semester = "Fall 2023"
                 student.add_course(course_code, course_name, semester)
                 sample_students.append(student)
                 return sample_students

class StudentGUI(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()
 
    def init_ui(self):
        self.tabs = QTabWidget()
 
        self.add_course_tab = self.create_add_course_tab()
        self.query_course_tab = self.create_query_course_tab()
        self.display_all_tab = self.create_display_all_tab()
 
        self.tabs.addTab(self.add_course_tab, "Add Course")
        self.tabs.addTab(self.query_course_tab, "Query Course")
        self.tabs.addTab(self.display_all_tab, "Display All Students")
 
        layout = QVBoxLayout()
        layout.addWidget(self.tabs)
        self.setLayout(layout)
 
        self.setGeometry(300, 300, 500, 400)
        self.setWindowTitle('STUST Student Information and Course Query Tool')
 
    def create_add_course_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()
 
        self.student_id_label_add = QLabel('Student ID:')
        self.student_id_edit_add = QLineEdit()
        self.name_label_add = QLabel('Name:')
        self.name_edit_add = QLineEdit()
 
        self.course_code_label_add = QLabel('Course Code:')
        self.course_code_edit_add = QLineEdit()
        self.course_name_label_add = QLabel('Course Name:')
        self.course_name_edit_add = QLineEdit()
        self.add_course_button_add = QPushButton('Add Course')
 
        self.add_course_button_add.clicked.connect(self.add_course)
 
        layout.addWidget(self.student_id_label_add)
        layout.addWidget(self.student_id_edit_add)
        layout.addWidget(self.name_label_add)
        layout.addWidget(self.name_edit_add)
        layout.addWidget(self.course_code_label_add)
        layout.addWidget(self.course_code_edit_add)
        layout.addWidget(self.course_name_label_add)
        layout.addWidget(self.course_name_edit_add)
        layout.addWidget(self.add_course_button_add)
 
        tab.setLayout(layout)
        return tab
 
    def create_query_course_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()
 
        self.student_id_label_query = QLabel('Student ID:')
        self.student_id_edit_query = QLineEdit()
        self.semester_label_query = QLabel('Semester:')
        self.semester_edit_query = QLineEdit()
        self.query_button_query = QPushButton('Query Courses')
        self.result_browser_query = QTextBrowser()
 
        self.query_button_query.clicked.connect(self.query_courses)
 
        layout.addWidget(self.student_id_label_query)
        layout.addWidget(self.student_id_edit_query)
        layout.addWidget(self.semester_label_query)
        layout.addWidget(self.semester_edit_query)
        layout.addWidget(self.query_button_query)
        layout.addWidget(self.result_browser_query)
 
        tab.setLayout(layout)
        return tab
 
    def create_display_all_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()
 
        self.query_button_display_all = QPushButton('Display All Students')
        self.result_browser_display_all = QTextBrowser()
 
        self.query_button_display_all.clicked.connect(self.display_all_students)
 
        layout.addWidget(self.query_button_display_all)
        layout.addWidget(self.result_browser_display_all)
 
        tab.setLayout(layout)
        return tab
 
    def add_course(self):
        student_id = self.student_id_edit_add.text()
        name = self.name_edit_add.text()
        course_code = self.course_code_edit_add.text()
        course_name = self.course_name_edit_add.text()
 
        # Check for missing input data
        if not student_id or not name or not course_code or not course_name:
            self.show_alert("Missing Data", "Please fill in all fields.")
            return
 
        # Create or load student profile
        student = Student(student_id, name)
        student.load_from_file(f"{student_id}_profile.json")
 
        semester = "Fall 2023"  # Set the default semester to Fall 2023 for simplicity
 
        # Add course to the student's profile
        student.add_course(course_code, course_name, semester)
 
        # Save student course information to the right JSON file
        student.save_to_file(f"{student_id}_profile.json")
 
        # Show success message box
        self.show_success_message("Course Added", "The course has been successfully added.")
 
    def query_courses(self):
        student_id = self.student_id_edit_query.text()
        semester = self.semester_edit_query.text()
 
        # Check for missing input data
        if not student_id or not semester:
            self.show_alert("Missing Data", "Please fill in all fields.")
            return
 
        # Create or load student profile
        student = Student(student_id, "")
        student.load_from_file(f"{student_id}_profile.json")
 
        # Query course information
        result_message, courses_info = student.query_course_info(semester)
 
        # Display result in the QTextBrowser
        self.result_browser_query.setPlainText(result_message)
        if courses_info:
            for course in courses_info:
                self.result_browser_query.append(f"{course['code']}: {course['name']}")
 
    def display_all_students(self):
        # Generate sample data for at least 10 students at STUST
        sample_students = generate_sample_data()
 
        # Save all student courses info to JSON file
        all_students_courses_info = []
        for student in sample_students:
            all_students_courses_info.extend(student.get_all_courses_info())
            # Save each student's course information to the right JSON file
            student.save_to_file(f"{student.student_id}_profile.json")
 
        # Generate and save Excel file
        self.generate_excel_file(all_students_courses_info)
 
        # Display all student course info in the QTextBrowser
        self.result_browser_display_all.clear()
        for info in all_students_courses_info:
            self.result_browser_display_all.append(f"{info['student_id']} - {info['name']}:")
            self.result_browser_display_all.append(f"  Semester: {info['semester']}")
            self.result_browser_display_all.append(f"  {info['course_code']}: {info['course_name']}")
 
    def generate_excel_file(self, all_students_courses_info):
        workbook = Workbook()
        worksheet = workbook.active
 
        # Set headers
        headers = ["Student ID", "Name", "Semester", "Course Code", "Course Name"]
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
 
        # Populate data
        for row_num, info in enumerate(all_students_courses_info, 2):
            worksheet.cell(row=row_num, column=1, value=info["student_id"])
            worksheet.cell(row=row_num, column=2, value=info["name"])
            worksheet.cell(row=row_num, column=3, value=info["semester"])
            worksheet.cell(row=row_num, column=4, value=info["course_code"])
            worksheet.cell(row=row_num, column=5, value=info["course_name"])
 
        # Save the Excel file
        workbook.save("students_courses_info_all.xlsx")
 
    def show_alert(self, title, message):
        alert = QMessageBox()
        alert.setWindowTitle(title)
        alert.setText(message)
        alert.exec()
 
    def show_success_message(self, title, message):
        success_message = QMessageBox()
        success_message.setWindowTitle(title)
        success_message.setText(message)
        success_message.exec()
 
def main():
    app = QApplication(sys.argv)
    gui = StudentGUI()
    gui.show()
    sys.exit(app.exec())
 
if __name__ == '__main__':
    main()